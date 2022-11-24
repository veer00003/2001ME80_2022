from datetime import datetime
start_time = datetime.now()

import math
# imported pandas library for accessing the input file
# then used shape to fetch the dimensions of pandas type object
import pandas as pd



def octant_analysis(mod, df):
	# AVERAGE

	# here used the mean() function for finding the average of U
	U_avg = df['U'].mean()
	# similarly here used mean function for V and W
	V_avg = df['V'].mean()
	W_avg = df['W'].mean()

	# made a column to store average of U
	df['U Avg'] = U_avg
	# and here given only 1st place of line to U avg otherwise average value will print in whole column
	df['U Avg'] = df['U Avg'].head(1)

	# similarly here made column for average of V and W
	# also here given only 1st place of line to W and V average otherwise average value will print in whole column
	df['V Avg'] = V_avg
	df['V Avg'] = df['V Avg'].head(1)

	df['W Avg'] = V_avg
	df['W Avg'] = df['W Avg'].head(1)

	# here defined the X,Y,Z and X=U' , Y=V' , Z=W'
	X = df['U'] - U_avg
	Y = df['V'] - V_avg
	Z = df['W'] - W_avg

	# here made the column for storing X and named the column as U'=U - U_avg and similarly done for Y and Z.
	df["U'=U - U_avg"] = X
	df["V'=V - V_avg"] = Y
	df["W'=W - W_avg"] = Z

	# FINDING OCTANT

	# here made the column for storing the value of octant
	df.insert(10, column="Octant", value="")
	# Octant=[]
	# using loop for taking the all values of the row and column
	for i in range(0, x):
		# for ease defined M, N, O
		M = df["U'=U - U_avg"][i]
		N = df["V'=V - V_avg"][i]
		O = df["W'=W - W_avg"][i]

	# finding octant value by using conditional statements if and elif

	# 1st quadrant(1) and positive z(+)
		if M > 0 and N > 0 and O > 0:
			print(1)
			df["Octant"][i] = 1

	# 1st quadrant(1) and negative z(-)
		elif M > 0 and N > 0 and O < 0:
			print(-1)
			df["Octant"][i] = -1

	# 2nd quadrant(2) and positive z(+)
		elif M < 0 and N > 0 and O > 0:
			print(2)
			df["Octant"][i] = 2

	# 2nd quadrant(2) and negative z(-)
		elif M < 0 and N > 0 and O < 0:
			print(-2)
			df["Octant"][i] = -2

	# 3rd quadrant(3) and positive z(+)
		elif M < 0 and N < 0 and O > 0:
			print(3)
			df["Octant"][i] = 3

	# 3rd quadrant(3) and negative z(-)
		elif M < 0 and N < 0 and O < 0:
			print(-3)
			df["Octant"][i] = -3

	# 4th quadrant(4) and positive z(+)
		elif M > 0 and N < 0 and O > 0:
			print(4)
			df["Octant"][i] = 4

	# 4th quadrant(4) and negative z(-)
		elif M > 0 and N < 0 and O < 0:
			print(-4)
			df["Octant"][i] = -4

			# COUNTING OCTANT

	# user input
	df.at[1, ''] = 'User Input'
	# creating row
	df.at[0, 'Octant ID'] = 'Overall Count'

	# counting the 1 and -1 putting list of Octant ID
	df.at[0, '1'] = list(df['Octant']).count(1)
	df.at[0, '-1'] = list(df['Octant']).count(-1)

	# counting the 1 and -1 putting list of Octant ID
	df.at[0, '2'] = list(df['Octant']).count(2)
	df.at[0, '-2'] = list(df['Octant']).count(-2)

	# counting the 1 and -1 putting list of Octant ID
	df.at[0, '3'] = list(df['Octant']).count(3)
	df.at[0, '-3'] = list(df['Octant']).count(-3)

	# counting the 1 and -1 putting list of Octant ID
	df.at[0, '4'] = list(df['Octant']).count(4)
	df.at[0, '-4'] = list(df['Octant']).count(-4)

	# COUNTING OCTANT VALUES FOR RANGES AND PROCESSING FOR FINAL SOLUTION

	# defining boundry for ranges
	boundary = []

	mod = 5000
	n = math.ceil(len(df.index)/mod)
	df.at[1, 'Octant ID'] = 'Mod {}'.format(mod)

	# using loop to devide all value of data in 6 ranges of 5000 difference
	for i in range(math.ceil(len(df.index)/mod)):
		boundary.append(i*mod)
	# appending the boundry and for last value taking the index length of df
	boundary.append(len(df.index))

	# countig the octants value in each range division by using loop
	for i in range(len(boundary) - 1):
		df.at[2 + i, 'Octant ID'] = str(boundary[i]) + \
			" to " + str(boundary[i + 1] - 1)

		df.at[2 + i, '1'] = df['Octant'].iloc[boundary[i]
			:(boundary[i+1])].value_counts()[1]
		df.at[2 + i, '-1'] = df['Octant'].iloc[boundary[i]
			:(boundary[i+1])].value_counts()[-1]

		df.at[2 + i, '2'] = df['Octant'].iloc[boundary[i]
			:(boundary[i+1])].value_counts()[2]
		df.at[2 + i, '-2'] = df['Octant'].iloc[boundary[i]
			:(boundary[i+1])].value_counts()[-2]

		df.at[2 + i, '3'] = df['Octant'].iloc[boundary[i]
			:(boundary[i+1])].value_counts()[3]
		df.at[2 + i, '-3'] = df['Octant'].iloc[boundary[i]
			:(boundary[i+1])].value_counts()[-3]

		df.at[2 + i, '4'] = df['Octant'].iloc[boundary[i]
			:(boundary[i+1])].value_counts()[4]
		df.at[2 + i, '-4'] = df['Octant'].iloc[boundary[i]
			:(boundary[i+1])].value_counts()[-4]

  # finding rank

	rank1 = []
	for j in range(0, n+2):
		Octant_name_ID_mapping = {'1': "Internal outward interaction", '-1': "External outward interaction", '2': "External Ejection",
								  '-2': "Internal Ejection", '3': "External inward interaction", '-3': "Internal inward interaction", '4': "Internal sweep", '-4': "External sweep"}

		Octant_count = []

		for i in ['1', '-1', '2', '-2', '3', '-3', '4', '-4']:
			Octant_count.append(df.at[j, i])
		Octant_count.sort(reverse=True)  # sorting overall count

		for i in ['1', '-1', '2', '-2', '3', '-3', '4', '-4']:
			for x in range(0, 8):
				if (Octant_count[x] == df.at[j, i]):
					# acessing indices after sorting for ranking
					df.at[j, "Rank("+i+")"] = x+1

		for k in ['1', '-1', '2', '-2', '3', '-3', '4', '-4']:
			if (df.loc[j, "Rank("+str(k)+")"] == 1):
				# Creating Rank1 Octant ID and Rank1 Octant name
				df.at[j, "Rank1 Octant ID"] = k
				df.at[j, "Rank1 Octant Name"] = Octant_name_ID_mapping[str(k)]
				if j == 0:
					continue
				else:
					# appending the rank 1 mod count for mod values
					rank1.append(int(k))

	df.at[n+5, '1'] = 'Octant Id'
	df.at[n+5, '-1'] = 'Octant Name'
	df.at[n+5, '2'] = 'Count of Rank 1 Mod Values'
	for j, x in enumerate([1, -1, 2, -2, 3, -3, 4, -4]):
		df.at[n+6+j, '1'] = x
		df.at[n+6+j, '-1'] = Octant_name_ID_mapping[str(x)]
		# printing the count of octant which is ranked 1
		df.at[n+6+j, '2'] = rank1.count(x)
	df = pd.concat([df.columns.to_frame().T, df], ignore_index=True)


	# TRANSITION

	df.at[13, 'Octant ID'] = 'Overall Transition Count'

	# These are the octant values
	Octant = ['1', '-1', '2', '-2', '3', '-3', '4', '-4']
	df.at[14, 1] = 'To'
	df.at[16, ''] = 'From'

	# Insert the table labels
	for i, x in enumerate(Octant):
		df.at[14, x] = x
		df.at[16 + i, 'Octant ID'] = x

	# Initialise table cells to zero
	for i in range(8):
		for j in range(8):
			df.at[16+i, Octant[j]] = 0

			x, y = 0, 0
	for i in range(boundary[0], boundary[-1]-1):
		from_octant = df["Octant"][i]
		to_octant = df["Octant"][i+1]
		x = Octant.index(str(from_octant))
		y = Octant.index(str(to_octant))

		df.at[16 + x, Octant[y]] += 1

		row_number = 28

	# The mod transition count tables are made for the boundaries
	for k in range(len(boundary)-1):
		df.at[row_number - 1,
			  'Octant ID'] = str(boundary[k]) + '-' + str(boundary[k + 1])
		df.at[row_number - 3, 'Octant ID'] = 'Mod Transition Count'
		df.at[row_number - 2, '1'] = 'To'
		df.at[row_number, ''] = 'From'
		for i, z in enumerate(Octant):
			df.at[row_number - 1, z] = z
			df.at[row_number + i, 'Octant ID'] = z
		for i in range(8):
			for j in range(8):
				df.at[row_number+i, Octant[j]] = 0
		for i in range(boundary[k], boundary[k+1]-1):
			from_octant = df["Octant"][i]
			to_octant = df["Octant"][i+1]
			x = Octant.index(str(from_octant))
			y = Octant.index(str(to_octant))

			df.at[row_number + x, Octant[y]] += 1

		row_number += 15


	Octant = [1, -1, 2, -2, 3, -3, 4, -4]
	for i, a in enumerate(Octant):
		df.at[i, 'octant'] = a
		c, maxC = 0, 0
		# Making subsequence of longest length
		for z in df['Octant']:
			if z == a:
				c += 1
			else:
				maxC = max(maxC, c)
				c = 0
		df.at[i, 'Longest Subsequence'] = maxC
		c, count = 0, 0

		# Making subsequence occurences  of longest length
		for z in df['Octant']:
			if z == a:
				c += 1
			else:
				if c == maxC:
					count += 1

				c = 0
		df.at[i, 'Count'] = count


	# making columns
	df['octn'] = ''
	df['longest subsequence length'] = ''
	df['count_'] = ''
	row = 0

	# finding longest subsequence length and range for 1
	df.at[row, 'octn'] = 1
	df.at[row, 'longest subsequence length'] = df.at[0, 'Longest Subsequence']
	df.at[row+1, 'octn'] = 'Time'
	df.at[row+1, 'longest subsequence length'] = 'From'
	df.at[row+1, 'count_'] = 'To'
	start = -1
	length = 0
	row += 2
	count = 0
	for i in range(len(df['Octant'])):
		if df.at[i, 'Octant'] == 1:
			length = 0
			start = i
			while (df.at[i, 'Octant'] == 1):
				i += 1
				length += 1
				if (i == len(df)):
					i -= 1
					break
			if length == df.at[0, 'Longest Subsequence']:
				df.at[row, 'longest subsequence length'] = df.at[start, 'Time']
				df.at[row, 'count_'] = df.at[i, 'Time']
				row += 1
				count += 1
			length = 0
	df.at[row-count-2, 'count_'] = count


	# finding longest subsequence length and range for -1
	df.at[row, 'octn'] = -1
	df.at[row, 'longest subsequence length'] = df.at[1, 'Longest Subsequence']
	df.at[row+1, 'octn'] = 'Time'
	df.at[row+1, 'longest subsequence length'] = 'From'
	df.at[row+1, 'count_'] = 'To'
	start = -1
	length = 0
	row += 2
	count = 0
	for i in range(len(df['Octant'])):
		if df.at[i, 'Octant'] == -1:
			length = 0
			start = i
			while (df.at[i, 'Octant'] == -1):
				i += 1
				length += 1
				if (i == len(df)):
					i -= 1
					break
			if length == df.at[1, 'Longest Subsequence']:
				df.at[row, 'longest subsequence length'] = df.at[start, 'Time']
				df.at[row, 'count_'] = df.at[i, 'Time']
				row += 1
				count += 1
			length = 0
	df.at[row-count-2, 'count_'] = count


	# finding longest subsequence length and range for 2
	df.at[row, 'octn'] = 2
	df.at[row, 'longest subsequence length'] = df.at[2, 'Longest Subsequence']
	df.at[row+1, 'octn'] = 'Time'
	df.at[row+1, 'longest subsequence length'] = 'From'
	df.at[row+1, 'count_'] = 'To'
	start = -1
	length = 0
	row += 2
	count = 0
	for i in range(len(df['Octant'])):
		if df.at[i, 'Octant'] == 2:
			length = 0
			start = i
			while (df.at[i, 'Octant'] == 2):
				i += 1
				length += 1
				if (i == len(df)):
					i -= 1
					break
			if length == df.at[2, 'Longest Subsequence']:
				df.at[row, 'longest subsequence length'] = df.at[start, 'Time']
				df.at[row, 'count_'] = df.at[i, 'Time']
				row += 1
				count += 1
			length = 0
	df.at[row-count-2, 'count_'] = count


	# finding longest subsequence length and range for -2
	df.at[row, 'octn'] = -2
	df.at[row, 'longest subsequence length'] = df.at[3, 'Longest Subsequence']
	df.at[row+1, 'octn'] = 'Time'
	df.at[row+1, 'longest subsequence length'] = 'From'
	df.at[row+1, 'count_'] = 'To'
	start = -1
	length = 0
	row += 2
	count = 0
	for i in range(len(df['Octant'])):
		if df.at[i, 'Octant'] == -2:
			length = 0
			start = i
			while (df.at[i, 'Octant'] == -2):
				i += 1
				length += 1
				if (i == len(df)):
					i -= 1
					break
			if length == df.at[3, 'Longest Subsequence']:
				df.at[row, 'longest subsequence length'] = df.at[start, 'Time']
				df.at[row, 'count_'] = df.at[i, 'Time']
				row += 1
				count += 1
			length = 0
	df.at[row-count-2, 'count_'] = count


	# finding longest subsequence length and range for 3
	df.at[row, 'octn'] = 3
	df.at[row, 'longest subsequence length'] = df.at[4, 'Longest Subsequence']
	df.at[row+1, 'octn'] = 'Time'
	df.at[row+1, 'longest subsequence length'] = 'From'
	df.at[row+1, 'count_'] = 'To'
	start = -1
	length = 0
	row += 2
	count = 0
	for i in range(len(df['Octant'])):
		if df.at[i, 'Octant'] == 3:
			length = 0
			start = i
			while (df.at[i, 'Octant'] == 3):
				i += 1
				length += 1
				if (i == len(df)):
					i -= 1
					break
			if length == df.at[4, 'Longest Subsequence']:
				df.at[row, 'longest subsequence length'] = df.at[start, 'Time']
				df.at[row, 'count_'] = df.at[i, 'Time']
				row += 1
				count += 1
			length = 0
	df.at[row-count-2, 'count_'] = count


	# finding longest subsequence length and range for -3
	df.at[row, 'octn'] = -3
	df.at[row, 'longest subsequence length'] = df.at[5, 'Longest Subsequence']
	df.at[row+1, 'octn'] = 'Time'
	df.at[row+1, 'longest subsequence length'] = 'From'
	df.at[row+1, 'count_'] = 'To'
	start = -1
	length = 0
	row += 2
	count = 0
	for i in range(len(df['Octant'])):
		if df.at[i, 'Octant'] == -3:
			length = 0
			start = i
			while (df.at[i, 'Octant'] == -3):
				i += 1
				length += 1
				if (i == len(df)):
					i -= 1
					break
			if length == df.at[5, 'Longest Subsequence']:
				df.at[row, 'longest subsequence length'] = df.at[start, 'Time']
				df.at[row, 'count_'] = df.at[i, 'Time']
				row += 1
				count += 1
			length = 0
	df.at[row-count-2, 'count_'] = count


	# finding longest subsequence length and range for 4
	df.at[row, 'octn'] = 4
	df.at[row, 'longest subsequence length'] = df.at[6, 'Longest Subsequence']
	df.at[row+1, 'octn'] = 'Time'
	df.at[row+1, 'longest subsequence length'] = 'From'
	df.at[row+1, 'count_'] = 'To'
	start = -1
	length = 0
	row += 2
	count = 0
	for i in range(len(df['Octant'])):
		if df.at[i, 'Octant'] == 4:
			length = 0
			start = i
			while (df.at[i, 'Octant'] == 4):
				i += 1
				length += 1
				if (i == len(df)):
					i -= 1
					break
			if length == df.at[6, 'Longest Subsequence']:
				df.at[row, 'longest subsequence length'] = df.at[start, 'Time']
				df.at[row, 'count_'] = df.at[i, 'Time']
				row += 1
				count += 1
			length = 0
	df.at[row-count-2, 'count_'] = count
	
	# finding longest subsequence length and range for -4
	df.at[row, 'octn'] = -4
	df.at[row, 'longest subsequence length'] = df.at[7, 'Longest Subsequence']
	df.at[row+1, 'octn'] = 'Time'
	df.at[row+1, 'longest subsequence length'] = 'From'
	df.at[row+1, 'count_'] = 'To'
	start = -1
	length = 0
	row += 2
	count = 0
	for i in range(len(df['Octant'])):
		if df.at[i, 'Octant'] == -4:
			length = 0
			start = i
			while (df.at[i, 'Octant'] == -4):
				i += 1
				length += 1
				if (i == len(df)):
					i -= 1
					break
			if length == df.at[7, 'Longest Subsequence']:
				df.at[row, 'longest subsequence length'] = df.at[start, 'Time']
				df.at[row, 'count_'] = df.at[i, 'Time']
				row += 1
				count += 1
			length = 0
	df.at[row-count-2, 'count_'] = count

from platform import python_version
import pandas as pd
import os
import math
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, colors, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


for filename in os.listdir('input'):
    f = os.path.join('input', filename)
    if os.path.isfile(f) and f.endswith('.xlsx'):
        df = pd.read_excel(f)  


        mod=5000

        octant_analysis(mod)
        df.to_excel(os.path.join('output', filename), index=None)

        file = load_workbook('output.xlsx')
        sheet = file.active

        side = Side(border_style='thin', color="000000")
        border = Border(top=side, bottom=side, left=side, right=side)
        for cell in sheet['AG1':'AI10']:
            for x in cell:
                x.border = border
        for cell in sheet['AK1':'AM' + str(sum([sheet['AI' + str(i)].value for i in range(3,11)]) + 9)]:
            for x in cell:
                x.border = border
        for cell in sheet['M1':'AE' + str((len(df) // mod) + 4)]:
            for x in cell:
                x.border = border
        for cell in sheet['N' + str((len(df) // mod) + 8):'P' + str((len(df) // mod) + 16)]:
            for x in cell:
                x.border = border
        for i in range((len(df) // mod) + 2):
            for cell in sheet['AO' + str(4 + 13*i) : 'AW' + str(12 + 13*i)]:
                for x in cell:
                    x.border = border

        style =DifferentialStyle(fill=PatternFill(bgColor='FFFF00'))
        rule = Rule(type="expression", dxf=style)
        rule.formula = ['V2=1']
        sheet.conditional_formatting.add("V2:AD"+str((len(df) // mod) + 4), rule)

        for i in range((len(df) // mod) + 2):
            rule2 = Rule(type='expression', dxf=style)
            rule2.formula = ['AP' + str(5 + 13*i) + '=MAX($AP' + str(5 + 13*i) + ':$AW' + str(5 + 13*i) + ')']
            sheet.conditional_formatting.add('AP' + str(5 + 13*i) + ':AW' + str(12 + 13*i), rule2)

        file.save(os.path.join('output', filename))
#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))