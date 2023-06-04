import streamlit as st
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side
from datetime import datetime
from io import BytesIO
import os
from tempfile import NamedTemporaryFile
from streamlit_option_menu import option_menu
from PIL import Image
start_time = datetime.now()



# ------------------------------------------


# mod value\

# mod=int(input("enter mod value"))
# mod=5000


# ----------------------------------------------------


regular = Side(border_style="medium", color="000000")
box = Border(top = regular,bottom=regular, left=regular,right=regular)


# ----------------------------------------------------



def Average(lst):
    return sum(lst) / len(lst)


def octant(u, v, w):     #tests the 8 diff conditions and returns the oct no
    if((u>=0) and (v>=0) and (w>=0)):
        return 1
    elif((u>=0) and (v>=0) and (w<0)):
        return -1
    elif((u<0) and (v>=0) and (w>=0)):
        return 2
    elif((u<0) and (v>=0) and (w<0)):
        return -2
    elif((u<0) and (v<0) and (w>=0)):
        return 3
    elif((u<0) and (v<0) and (w<0)):
        return -3
    elif((u>=0) and (v<0) and (w>=0)):
        return 4
    else:
        return -4


def octname (x):
    octname_dict = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

    if x == 1:
        return (octname_dict['1'])

    elif x == -1:
        return (octname_dict['-1'])

    elif x == 2:
        return (octname_dict['2'])

    elif x == -2:
        return (octname_dict['-2'])

    elif x == 3:
        return (octname_dict['3'])

    elif x == -3:
        return (octname_dict['-3'])

    elif x == 4:
        return (octname_dict['4'])

    elif x == -4:
        return (octname_dict['-4'])



def maxvalueranker (list, ranklist):

    for x in list:


        ranklist.append(len(list)-(sorted(list).index(x)))

        # since the sorted list is ascending in order,
        # we subtract the indices from the length to get our ranks


    for x in list:

        # now appending the value with rank 1


        # (id probably use a list to use a list someday)


        if (len(list)-(sorted(list).index(x))) == 1:

            if (list.index(x)) == 0:
                ranklist.append(1)

            elif (list.index(x)) == 1:
                ranklist.append(-1)

            elif (list.index(x)) == 2:
                ranklist.append(2)

            elif (list.index(x)) == 3:
                ranklist.append(-2)

            elif (list.index(x)) == 4:
                ranklist.append(3)

            elif (list.index(x)) == 5:
                ranklist.append(-3)

            elif (list.index(x)) == 6:
                ranklist.append(4)

            elif (list.index(x)) == 7:
                ranklist.append(-4)





def transcounter(oct, matrix):
    i=0
    for i in range (8):
        matrix.append([])
        matrix[i]=[0,0,0,0,0,0,0,0]

    #print(matrix)

    L=len(oct)


    # basically what this entire piece of code below does is
    # one-by-one compare each element of the ocant column
    # and the element below it and checks the transition.
    #
    # depending on the transition, it keeps appending 1 to
    # a matrix which looks the same as the table we need.

    k=0
    i=0
    while k<(L-1):
        #print("mainwhile")




        if oct[k]==1:
            #print("mainif")
            if oct[k+1]==1:
                matrix[0][0]=matrix[0][0]+1
                #print("if1")
                k=k+1
                continue

            elif oct[k+1]==-1:
                matrix[0][1]=matrix[0][1]+1
                #print("if2")
                k=k+1
                continue

            elif oct[k+1]==2:
                matrix[0][2]=matrix[0][2]+1
                #print("if3")
                k=k+1
                continue

            elif oct[k+1]==-2:
                matrix[0][3]=matrix[0][3]+1
                #print("if4")
                k=k+1
                continue

            elif oct[k+1]==3:
                matrix[0][4]=matrix[0][4]+1
                #print("if5")
                k=k+1
                continue

            elif oct[k+1]==-3:
                matrix[0][5]=matrix[0][5]+1
                #print("if6")
                k=k+1
                continue

            elif oct[k+1]==4:
                matrix[0][6]=matrix[0][6]+1
                #print("if7")
                k=k+1
                continue

            elif oct[k+1]==-4:
                matrix[0][7]=matrix[0][7]+1
                #print("if8")
                k=k+1
                continue





        elif oct[k]==-1:
            #print("mainif")
            if oct[k+1]==1:
                matrix[1][0]=matrix[1][0]+1
                #print("if1")
                k=k+1
                continue

            elif oct[k+1]==-1:
                matrix[1][1]=matrix[1][1]+1
                #print("if2")
                k=k+1
                continue

            elif oct[k+1]==2:
                matrix[1][2]=matrix[1][2]+1
                #print("if3")
                k=k+1
                continue

            elif oct[k+1]==-2:
                matrix[1][3]=matrix[1][3]+1
                #print("if4")
                k=k+1
                continue

            elif oct[k+1]==3:
                matrix[1][4]=matrix[1][4]+1
                #print("if5")
                k=k+1
                continue

            elif oct[k+1]==-3:
                matrix[1][5]=matrix[1][5]+1
                #print("if6")
                k=k+1
                continue

            elif oct[k+1]==4:
                matrix[1][6]=matrix[1][6]+1
                #print("if7")
                k=k+1
                continue

            elif oct[k+1]==-4:
                matrix[1][7]=matrix[1][7]+1
                #print("if8")
                k=k+1
                continue





        elif oct[k]==2:
            #print("mainif")
            if oct[k+1]==1:
                matrix[2][0]=matrix[2][0]+1
                #print("if1")
                k=k+1
                continue

            elif oct[k+1]==-1:
                matrix[2][1]=matrix[2][1]+1
                #print("if2")
                k=k+1
                continue

            elif oct[k+1]==2:
                matrix[2][2]=matrix[2][2]+1
                #print("if3")
                k=k+1
                continue

            elif oct[k+1]==-2:
                matrix[2][3]=matrix[2][3]+1
                #print("if4")
                k=k+1
                continue

            elif oct[k+1]==3:
                matrix[2][4]=matrix[2][4]+1
                #print("if5")
                k=k+1
                continue

            elif oct[k+1]==-3:
                matrix[2][5]=matrix[2][5]+1
                #print("if6")
                k=k+1
                continue

            elif oct[k+1]==4:
                matrix[2][6]=matrix[2][6]+1
                #print("if7")
                k=k+1
                continue

            elif oct[k+1]==-4:
                matrix[2][7]=matrix[2][7]+1
                #print("if8")
                k=k+1
                continue





        elif oct[k]==-2:
            #print("mainif")
            if oct[k+1]==1:
                matrix[3][0]=matrix[3][0]+1
                #print("if1")
                k=k+1
                continue

            elif oct[k+1]==-1:
                matrix[3][1]=matrix[3][1]+1
                #print("if2")
                k=k+1
                continue

            elif oct[k+1]==2:
                matrix[3][2]=matrix[3][2]+1
                #print("if3")
                k=k+1
                continue

            elif oct[k+1]==-2:
                matrix[3][3]=matrix[3][3]+1
                #print("if4")
                k=k+1
                continue

            elif oct[k+1]==3:
                matrix[3][4]=matrix[3][4]+1
                #print("if5")
                k=k+1
                continue

            elif oct[k+1]==-3:
                matrix[3][5]=matrix[3][5]+1
                #print("if6")
                k=k+1
                continue

            elif oct[k+1]==4:
                matrix[3][6]=matrix[3][6]+1
                #print("if7")
                k=k+1
                continue

            elif oct[k+1]==-4:
                matrix[3][7]=matrix[3][7]+1
                #print("if8")
                k=k+1
                continue




        elif oct[k]==3:
            #print("mainif")
            if oct[k+1]==1:
                matrix[4][0]=matrix[4][0]+1
                #print("if1")
                k=k+1
                continue

            elif oct[k+1]==-1:
                matrix[4][1]=matrix[4][1]+1
                #print("if2")
                k=k+1
                continue

            elif oct[k+1]==2:
                matrix[4][2]=matrix[4][2]+1
                #print("if3")
                k=k+1
                continue

            elif oct[k+1]==-2:
                matrix[4][3]=matrix[4][3]+1
                #print("if4")
                k=k+1
                continue

            elif oct[k+1]==3:
                matrix[4][4]=matrix[4][4]+1
                #print("if5")
                k=k+1
                continue

            elif oct[k+1]==-3:
                matrix[4][5]=matrix[4][5]+1
                #print("if6")
                k=k+1
                continue

            elif oct[k+1]==4:
                matrix[4][6]=matrix[4][6]+1
                #print("if7")
                k=k+1
                continue

            elif oct[k+1]==-4:
                matrix[4][7]=matrix[4][7]+1
                #print("if8")
                k=k+1
                continue




        elif oct[k]==-3:
            #print("mainif")
            if oct[k+1]==1:
                matrix[5][0]=matrix[5][0]+1
                #print("if1")
                k=k+1
                continue

            elif oct[k+1]==-1:
                matrix[5][1]=matrix[5][1]+1
                #print("if2")
                k=k+1
                continue

            elif oct[k+1]==2:
                matrix[5][2]=matrix[5][2]+1
                #print("if3")
                k=k+1
                continue

            elif oct[k+1]==-2:
                matrix[5][3]=matrix[5][3]+1
                #print("if4")
                k=k+1
                continue

            elif oct[k+1]==3:
                matrix[5][4]=matrix[5][4]+1
                #print("if5")
                k=k+1
                continue

            elif oct[k+1]==-3:
                matrix[5][5]=matrix[5][5]+1
                #print("if6")
                k=k+1
                continue

            elif oct[k+1]==4:
                matrix[5][6]=matrix[5][6]+1
                #print("if7")
                k=k+1
                continue

            elif oct[k+1]==-4:
                matrix[5][7]=matrix[5][7]+1
                #print("if8")
                k=k+1
                continue




        elif oct[k]==4:
            #print("mainif")
            if oct[k+1]==1:
                matrix[6][0]=matrix[6][0]+1
                #print("if1")
                k=k+1
                continue

            elif oct[k+1]==-1:
                matrix[6][1]=matrix[6][1]+1
                #print("if2")
                k=k+1
                continue

            elif oct[k+1]==2:
                matrix[6][2]=matrix[6][2]+1
                #print("if3")
                k=k+1
                continue

            elif oct[k+1]==-2:
                matrix[6][3]=matrix[6][3]+1
                #print("if4")
                k=k+1
                continue

            elif oct[k+1]==3:
                matrix[6][4]=matrix[6][4]+1
                #print("if5")
                k=k+1
                continue

            elif oct[k+1]==-3:
                matrix[6][5]=matrix[6][5]+1
                #print("if6")
                k=k+1
                continue

            elif oct[k+1]==4:
                matrix[6][6]=matrix[6][6]+1
                #print("if7")
                k=k+1
                continue

            elif oct[k+1]==-4:
                matrix[6][7]=matrix[6][7]+1
                #print("if8")
                k=k+1
                continue




        elif oct[k]==-4:
            #print("mainif")
            if oct[k+1]==1:
                matrix[7][0]=matrix[7][0]+1
                #print("if1")
                k=k+1
                continue

            elif oct[k+1]==-1:
                matrix[7][1]=matrix[7][1]+1
                #print("if2")
                k=k+1
                continue

            elif oct[k+1]==2:
                matrix[7][2]=matrix[7][2]+1
                #print("if3")
                k=k+1
                continue

            elif oct[k+1]==-2:
                matrix[7][3]=matrix[7][3]+1
                #print("if4")
                k=k+1
                continue

            elif oct[k+1]==3:
                matrix[7][4]=matrix[7][4]+1
                #print("if5")
                k=k+1
                continue

            elif oct[k+1]==-3:
                matrix[7][5]=matrix[7][5]+1
                #print("if6")
                k=k+1
                continue

            elif oct[k+1]==4:
                matrix[7][6]=matrix[7][6]+1
                #print("if7")
                k=k+1
                continue

            elif oct[k+1]==-4:
                matrix[7][7]=matrix[7][7]+1
                #print("if8")
                k=k+1
                continue




        else:
            print("aidal ki kori dila tumi")
            k=k+1




def octant_longest_subsequence_count_with_range(octlist, list, countlist, time, timecnt):



    # modifications from the previous code here:
    # we take in the time list and have the new shiny timecnt 2d list to store
    # the initial and final times for each octant thingy.



    # okay okay,
    # what we do here is have two lists:
    # the one by the name of 'list' stores the number of elements in the
    # longest subsequence, and the 'countlist' stores the number of times
    # that pattern occurs.
    #
    # now since our entire algorithm is based on manipulating lists,
    # stuff does get pretty easy as everything is just list manipulation.



    k=0

    # these variables go in the 'list' list

    subcnt1=1
    subcntm1=1
    subcnt2=1
    subcntm2=1
    subcnt3=1
    subcntm3=1
    subcnt4=1
    subcntm4=1

    # and these go in the 'countlist' list

    occurrencecounter1=0
    occurrencecounterm1=0
    occurrencecounter2=0
    occurrencecounterm2=0
    occurrencecounter3=0
    occurrencecounterm3=0
    occurrencecounter4=0
    occurrencecounterm4=0

    L=len(octlist)
    ti=0
    tf=0

    # we iterate the thing for the entire oct list

    while k<(L-1):

        if octlist[k]==1:


            # --------------------------------------------------
            # this little neat bit here stores the time when a
            # subsequence is started


            if (octlist[k]!=octlist[k-1]) or (k==0):   #new addition
                ti=time[k]                             #new addition




            # --------------------------------------------------





            if octlist[k]==octlist[k+1]:
                subcnt1=subcnt1+1


            elif (octlist[k]==octlist[k-1]) and (subcnt1==list[0]):  #gotta initialise the list with 1

                # when the iteration hits a value which is the last in the line of
                # the sequence of numbers, if the count of numbers is equal to the
                # count of the biggest subsequence already stored, occurrence
                # counter is incremented and subcount is set to 1

                occurrencecounter1=occurrencecounter1+1
                #list[0]=subcnt1
                subcnt1=1


                # and this bit here stores the final time and appends them
                # to the 2d list timecnt.
                #
                # timecnt has the following structure:
                # [[tstart1, tend1, tstart2, tend2],   [],   [],   [].....]
                #          ^
                #          |
                #          for +1


                tf=time[k]               #new addition
                timecnt[0].append(ti)    #new addition
                timecnt[0].append(tf)    #new addition


            elif (octlist[k]==octlist[k-1]) and (subcnt1>list[0]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                # this here is to save the size of the subseq to the list when it
                # hits the last element of the subseq

                occurrencecounter1=1
                list[0]=subcnt1
                # print("there ", subcnt1)
                subcnt1=1



                #first erase whatever time data there already is in there

                # if a bigger subsequence is found, it clears the exisiting
                # info

                timecnt[0].clear()       #new addition

                tf=time[k]               #new addition
                timecnt[0].append(ti)    #new addition
                timecnt[0].append(tf)    #new addition


            elif subcnt1==list[0]:
                occurrencecounter1=occurrencecounter1+1
                subcnt1=1

                tf=time[k]               #new addition
                timecnt[0].append(ti)    #new addition
                timecnt[0].append(tf)    #new addition

                # and this is what happens to single element subsequences

            else:
                subcnt1=1



        # aaaaaaaaaaand repeat!
        # (dont mind the random comments in between there are too many to remove)


        if octlist[k]==-1:




            if (octlist[k]!=octlist[k-1]) or (k==0):   #new addition
                ti=time[k]                             #new addition







            if octlist[k]==octlist[k+1]:
                subcntm1=subcntm1+1


            elif (octlist[k]==octlist[k-1]) and (subcntm1==list[1]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounterm1=occurrencecounterm1+1
                #list[0]=subcnt1
                subcntm1=1


                tf=time[k]               #new addition
                timecnt[1].append(ti)    #new addition
                timecnt[1].append(tf)    #new addition


            elif (octlist[k]==octlist[k-1]) and (subcntm1>list[1]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounterm1=1
                list[1]=subcntm1
                subcntm1=1


                #first erase whatever time data there already is in there

                timecnt[1].clear()       #new addition

                tf=time[k]               #new addition
                timecnt[1].append(ti)    #new addition
                timecnt[1].append(tf)    #new addition

            elif subcntm1==list[1]:
                occurrencecounterm1=occurrencecounterm1+1
                subcntm1=1

                tf=time[k]               #new addition
                timecnt[1].append(ti)    #new addition
                timecnt[1].append(tf)    #new addition

            else:
                subcntm1=1





        if octlist[k]==2:

            if (octlist[k]!=octlist[k-1]) or (k==0):   #new addition
                ti=time[k]                             #new addition

            if octlist[k]==octlist[k+1]:
                subcnt2=subcnt2+1


            elif (octlist[k]==octlist[k-1]) and (subcnt2==list[2]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounter2=occurrencecounter2+1
                #list[0]=subcnt1
                subcnt2=1

                tf=time[k]               #new addition
                timecnt[2].append(ti)    #new addition
                timecnt[2].append(tf)    #new addition


            elif (octlist[k]==octlist[k-1]) and (subcnt2>list[2]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounter2=1
                list[2]=subcnt2
                subcnt2=1


                #first erase whatever time data there already is in there

                timecnt[2].clear()       #new addition

                tf=time[k]               #new addition
                timecnt[2].append(ti)    #new addition
                timecnt[2].append(tf)    #new addition

            elif subcnt2==list[2]:
                occurrencecounter2=occurrencecounter2+1
                subcnt2=1

                tf=time[k]               #new addition
                timecnt[2].append(ti)    #new addition
                timecnt[2].append(tf)    #new addition

            else:
                subcnt2=1




        if octlist[k]==-2:

            if (octlist[k]!=octlist[k-1]) or (k==0):   #new addition
                ti=time[k]                             #new addition

            if octlist[k]==octlist[k+1]:
                subcntm2=subcntm2+1



            elif (octlist[k]==octlist[k-1]) and (subcntm2==list[3]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounterm2=occurrencecounterm2+1
                #list[0]=subcnt1
                subcntm2=1

                tf=time[k]               #new addition
                timecnt[3].append(ti)    #new addition
                timecnt[3].append(tf)    #new addition


            elif (octlist[k]==octlist[k-1]) and (subcntm2>list[3]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounterm2=1
                list[3]=subcntm2
                subcntm2=1


                #first erase whatever time data there already is in there

                timecnt[3].clear()       #new addition

                tf=time[k]               #new addition
                timecnt[3].append(ti)    #new addition
                timecnt[3].append(tf)    #new addition

            elif subcntm2==list[3]:
                occurrencecounterm2=occurrencecounterm2+1
                subcntm2=1

                tf=time[k]               #new addition
                timecnt[3].append(ti)    #new addition
                timecnt[3].append(tf)    #new addition

            else:
                subcntm2=1





        if octlist[k]==3:

            if (octlist[k]!=octlist[k-1]) or (k==0):   #new addition
                ti=time[k]                             #new addition

            if octlist[k]==octlist[k+1]:
                subcnt3=subcnt3+1


            elif (octlist[k]==octlist[k-1]) and (subcnt3==list[4]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounter3=occurrencecounter3+1
                #list[0]=subcnt1
                subcnt3=1

                tf=time[k]               #new addition
                timecnt[4].append(ti)    #new addition
                timecnt[4].append(tf)    #new addition


            elif (octlist[k]==octlist[k-1]) and (subcnt3>list[4]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounter3=1
                list[4]=subcnt3
                subcnt3=1


                #first erase whatever time data there already is in there

                timecnt[4].clear()       #new addition

                tf=time[k]               #new addition
                timecnt[4].append(ti)    #new addition
                timecnt[4].append(tf)    #new addition

            elif subcnt3==list[4]:
                occurrencecounter3=occurrencecounter3+1
                subcnt3=1

                tf=time[k]               #new addition
                timecnt[4].append(ti)    #new addition
                timecnt[4].append(tf)    #new addition

            else:
                subcnt3=1






        if octlist[k]==-3:

            if (octlist[k]!=octlist[k-1]) or (k==0):   #new addition
                ti=time[k]                             #new addition

            if octlist[k]==octlist[k+1]:
                subcntm3=subcntm3+1


            elif (octlist[k]==octlist[k-1]) and (subcntm3==list[5]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounterm3=occurrencecounterm3+1
                #list[0]=subcnt1
                subcntm3=1

                tf=time[k]               #new addition
                timecnt[5].append(ti)    #new addition
                timecnt[5].append(tf)    #new addition


            elif (octlist[k]==octlist[k-1]) and (subcntm3>list[5]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounterm3=1
                list[5]=subcntm3
                subcntm3=1


                #first erase whatever time data there already is in there

                timecnt[5].clear()       #new addition

                tf=time[k]               #new addition
                timecnt[5].append(ti)    #new addition
                timecnt[5].append(tf)    #new addition

            elif subcntm3==list[5]:
                occurrencecounterm3=occurrencecounterm3+1
                subcntm3=1

                tf=time[k]               #new addition
                timecnt[5].append(ti)    #new addition
                timecnt[5].append(tf)    #new addition

            else:
                subcntm3=1





        if octlist[k]==4:

            if (octlist[k]!=octlist[k-1]) or (k==0):   #new addition
                ti=time[k]                             #new addition

            if octlist[k]==octlist[k+1]:
                subcnt4=subcnt4+1


            elif (octlist[k]==octlist[k-1]) and (subcnt4==list[6]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounter4=occurrencecounter4+1
                #list[0]=subcnt1
                subcnt4=1

                tf=time[k]               #new addition
                timecnt[6].append(ti)    #new addition
                timecnt[6].append(tf)    #new addition


            elif (octlist[k]==octlist[k-1]) and (subcnt4>list[6]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounter4=1
                list[6]=subcnt4
                subcnt4=1


                #first erase whatever time data there already is in there

                timecnt[6].clear()       #new addition

                tf=time[k]               #new addition
                timecnt[6].append(ti)    #new addition
                timecnt[6].append(tf)    #new addition

            elif subcnt4==list[6]:
                occurrencecounter4=occurrencecounter4+1
                subcnt4=1

                tf=time[k]               #new addition
                timecnt[6].append(ti)    #new addition
                timecnt[6].append(tf)    #new addition

            else:
                subcnt4=1





        if octlist[k]==-4:

            if (octlist[k]!=octlist[k-1]) or (k==0):   #new addition
                ti=time[k]                             #new addition

            if octlist[k]==octlist[k+1]:
                subcntm4=subcntm4+1


            elif (octlist[k]==octlist[k-1]) and (subcntm4==list[7]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounterm4=occurrencecounterm4+1
                #list[0]=subcnt1
                subcntm4=1

                tf=time[k]               #new addition
                timecnt[7].append(ti)    #new addition
                timecnt[7].append(tf)    #new addition


            elif (octlist[k]==octlist[k-1]) and (subcntm4>list[7]):  #gotta initialise the list with 1

                #something here to save the number above ughh

                occurrencecounterm4=1
                list[7]=subcntm4
                subcntm4=1


                #first erase whatever time data there already is in there

                timecnt[7].clear()       #new addition

                tf=time[k]               #new addition
                timecnt[7].append(ti)    #new addition
                timecnt[7].append(tf)    #new addition

            elif subcntm4==list[7]:
                occurrencecounterm4=occurrencecounterm4+1
                subcntm4=1

                tf=time[k]               #new addition
                timecnt[7].append(ti)    #new addition
                timecnt[7].append(tf)    #new addition

            else:
                subcntm4=1




        k=k+1


    countlist[0]=occurrencecounter1
    countlist[1]=occurrencecounterm1
    countlist[2]=occurrencecounter2
    countlist[3]=occurrencecounterm2
    countlist[4]=occurrencecounter3
    countlist[5]=occurrencecounterm3
    countlist[6]=occurrencecounter4
    countlist[7]=occurrencecounterm4








def mainprog(inputfile, mod, buttonnumber, option, pathlength):

	# inputfilename='tut07/input/'+inputfile+'.xlsx'

	inputfilename=inputfile

	# print('AUGHHHHHHHH')
	# print(inputfilename)
	#
	#
	# wb = load_workbook(inputfilename)
	# sheet = wb["Sheet1"]

	try:
		wb = load_workbook(inputfilename)
		sheet = wb["Sheet1"]
	except:
		print("check the path maybe")
		print("adios")
		exit()



	# yep, one list per column here


	t=[]
	u=[]
	v=[]
	w=[]
	#i=0
	#j=0
	row=1
	row_count = sheet.max_row
	for row in range (2, row_count+1):
	    # if j>0:
	    #
	    #
	    # j=j+1

		cellt='A'+str(row)
		cellu='B'+str(row)
		cellv='C'+str(row)
		cellw='D'+str(row)

		at=sheet[cellt]
		bt=sheet[cellu]
		ct=sheet[cellv]
		dt=sheet[cellw]

		a=at.value
		b=bt.value
		c=ct.value
		d=dt.value

		#print (a)

		# traversing the columns and appending the values at the end of
		# the lists we just created

		t.append(float(a))
		u.append(float(b))
		v.append(float(c))
		w.append(float(d))

	    #i=i+1



	av_u=Average(u)
	av_v=Average(v)
	av_w=Average(w)
	#ind=3
	#print (u[ind])

	# print (av_u, av_v, av_w)


	#ok this works
	#time for another list loop ugh


	#this gives the u-uavg
	#yes more lists (totally flexing the memory usage)

	U=[]
	V=[]
	W=[]

	L=len(u)
	k=0
	while k<L:
		A=u[k]-av_u
		B=v[k]-av_v
		C=w[k]-av_w
		U.append(A)
		V.append(B)
		W.append(C)
		k=k+1



	#print (U[0], V[0], W[0])

	#AUGHHHHHHHH


	#time for the quadrant thingy (octant actually im sorry)
	oct=[]
	K=0
	while K<L:
		e1=U[K]
		e2=V[K]
		e3=W[K]
		q=octant(e1, e2, e3)   #goes way up there
		oct.append(q)
		K=K+1
	#print(oct[0:5000].count(1))



	#IT WORKSSSSS!



	#counting the stuff
	c1=oct.count(1)
	cm1=oct.count(-1)
	c2=oct.count(2)
	cm2=oct.count(-2)
	c3=oct.count(3)
	cm3=oct.count(-3)
	c4=oct.count(4)
	cm4=oct.count(-4)

	# now that i tink about it, it seems like a better idea to have these
	# in a list


	#print(c1, c2, c3, c4)
	# print (L)



	if mod>L:
		print("max value of mod has exceeded sample size") #safetycheck
		#return
	else:
		ind=mod
		cmod = []
		rankmatrix=[]

		ran=(L//mod)+1



		#basically, cant do 2d arrays like c so appending blank lists to the list
		#and storing the calculated stuff in there


		# cmod structure:
		#
		# [[counts of +1], [counts of -1], [blah], [blah]]
		# or,
		# [[column], [column],....]

		for i in range (8):
			cmod.append([])
			rankmatrix.append([])

		x=0   #generous usage of variables here (rip efficiency)
		i=0
		while x<=L:
			cmod[i].append(oct[x:x+mod].count(1))
			x=x+mod

		x=0
		i=1
		while x<=L:
			cmod[i].append(oct[x:x+mod].count(-1))
			x=x+mod

		x=0
		i=2
		while x<=L:
			cmod[i].append(oct[x:x+mod].count(2))
			x=x+mod

		x=0
		i=3
		while x<=L:
			cmod[i].append(oct[x:x+mod].count(-2))
			x=x+mod

		x=0
		i=4
		while x<=L:
			cmod[i].append(oct[x:x+mod].count(3))
			x=x+mod

		x=0
		i=5
		while x<=L:
			cmod[i].append(oct[x:x+mod].count(-3))
			x=x+mod

		x=0
		i=6
		while x<=L:
			cmod[i].append(oct[x:x+mod].count(4))
			x=x+mod

		x=0
		i=7
		while x<=L:
			cmod[i].append(oct[x:x+mod].count(-4))
			x=x+mod

		#print(cmod)





	overall_transcount_matrix=[]  # this one for the overall transition counts
	transcounter(oct, overall_transcount_matrix)

	#print (overall_transcount_matrix)


	x=0
	c=0


	# for the next part, that is, the matrices of the restricted octant columns,
	# we will have a list of matrices
	# [[matrix1], [matrix2], .......]
	# the structure of [matrix1], [matrix2], etc are exactly the same as
	# the overall transition count matrix.

	modtransmatrixlist=[]
	while x<=L:
		templist=oct[x:x+mod]
		#print(templist)


		tempmatrixlist=[]
		transcounter(templist, tempmatrixlist)

		modtransmatrixlist.append(tempmatrixlist)
		#keep appending the mod transition matrices as they are formed
		#into the 3d array



		x=x+mod





	seqlength=[1,1,1,1,1,1,1,1]
	occount=[1,1,1,1,1,1,1,1,]
	timecounter=[[],[],[],[],[],[],[],[]]

	octant_longest_subsequence_count_with_range(oct, seqlength, occount, t, timecounter)
	# print(seqlength)
	# print(occount)
	# print(oct)
	print(timecounter)










	# ---------------------------
	# NOW WE WRITE
	# ---------------------------


	try:
		wb2=Workbook()
		Sheet1=wb2.active

	except:
		print("for some reason, cant open a workbook/worksheet")
		print("terminating, adios")
		exit()



	#the headers






	Sheet1['A2']='Time'
	Sheet1['B2']='U'
	Sheet1['C2']='V'
	Sheet1['D2']='W'
	Sheet1['E2']='Uavg'
	Sheet1['F2']='Vavg'
	Sheet1['G2']='Wavg'
	Sheet1['H2']='U\'=U-Uavg'
	Sheet1['I2']='V\'=V-Vavg'
	Sheet1['J2']='W\'=W-Wavg'
	Sheet1['K2']='Octant'

	Sheet1['E3']=av_u
	Sheet1['F3']=av_v
	Sheet1['G3']=av_w

	#the simple columns

	Sheet1['AS3']='octant ##'
	Sheet1['AS4']='+1'
	Sheet1['AS5']='-1'
	Sheet1['AS6']='+2'
	Sheet1['AS7']='-2'
	Sheet1['AS8']='+3'
	Sheet1['AS9']='-3'
	Sheet1['AS10']='+4'
	Sheet1['AS11']='-4'


	Sheet1['AT3']='longest subsequence length'
	Sheet1['AT4']=seqlength[0]
	Sheet1['AT5']=seqlength[1]
	Sheet1['AT6']=seqlength[2]
	Sheet1['AT7']=seqlength[3]
	Sheet1['AT8']=seqlength[4]
	Sheet1['AT9']=seqlength[5]
	Sheet1['AT10']=seqlength[6]
	Sheet1['AT11']=seqlength[7]


	Sheet1['AU3']='count'
	Sheet1['AU4']=occount[0]
	Sheet1['AU5']=occount[1]
	Sheet1['AU6']=occount[2]
	Sheet1['AU7']=occount[3]
	Sheet1['AU8']=occount[4]
	Sheet1['AU9']=occount[5]
	Sheet1['AU10']=occount[6]
	Sheet1['AU11']=occount[7]

	for x in range(9):
		for c in Sheet1[f'AS{x+3}':f'AU{x+3}'][0]:
			c.border = box










	#lets get the big columns printed

	row=3
	index=0
	row_count = sheet.max_row
	for row in range (3, row_count+1):


		cellt='A'+str(row)
		cellu='B'+str(row)
		cellv='C'+str(row)
		cellw='D'+str(row)

		cellU='H'+str(row)
		cellV='I'+str(row)
		cellW='J'+str(row)
		celloct='K'+str(row)

		Sheet1[cellt]=t[index]
		Sheet1[cellu]=u[index]
		Sheet1[cellv]=v[index]
		Sheet1[cellw]=w[index]

		Sheet1[cellt].number_format='0.00'
		Sheet1[cellu].number_format='0.00'
		Sheet1[cellv].number_format='0.00'
		Sheet1[cellw].number_format='0.00'

		Sheet1[cellU]=U[index]
		Sheet1[cellV]=V[index]
		Sheet1[cellW]=W[index]
		Sheet1[celloct]=oct[index]

		Sheet1[cellU].number_format='0.00'
		Sheet1[cellV].number_format='0.00'
		Sheet1[cellW].number_format='0.00'

		index=index+1

	#that worked

	#the stuff above the mod table cums here
	Sheet1['N2']='+1'
	Sheet1['O2']='-1'
	Sheet1['P2']='+2'
	Sheet1['Q2']='-2'
	Sheet1['R2']='+3'
	Sheet1['S2']='-3'
	Sheet1['T2']='+4'
	Sheet1['U2']='-4'

	# Sheet1['V1']='+1'
	# Sheet1['W1']='-1'
	# Sheet1['X1']='+2'
	# Sheet1['Y1']='-2'
	# Sheet1['Z1']='+3'
	# Sheet1['AA1']='-3'
	# Sheet1['AB1']='+4'
	# Sheet1['AC1']='-4'

	Sheet1['V2']='Rank octant 1'
	Sheet1['W2']='Rank octant -1'
	Sheet1['X2']='Rank octant 2'
	Sheet1['Y2']='Rank octant -2'
	Sheet1['Z2']='Rank octant 3'
	Sheet1['AA2']='Rank octant -3'
	Sheet1['AB2']='Rank octant 4'
	Sheet1['AC2']='Rank octant -4'
	Sheet1['AD2']='Rank 1 octant ID'
	Sheet1['AE2']='Rank 1 octant name'

	Sheet1['M3']='Overall Count'

	Sheet1['N3']=c1
	Sheet1['O3']=cm1
	Sheet1['P3']=c2
	Sheet1['Q3']=cm2
	Sheet1['R3']=c3
	Sheet1['S3']=cm3
	Sheet1['T3']=c4
	Sheet1['U3']=cm4




	cmlist=[c1, cm1, c2, cm2, c3, cm3, c4, cm4]

	overallrank=[]

	maxvalueranker(cmlist, overallrank)


	Sheet1['V3']=overallrank[0]
	Sheet1['W3']=overallrank[1]
	Sheet1['X3']=overallrank[2]
	Sheet1['Y3']=overallrank[3]
	Sheet1['Z3']=overallrank[4]
	Sheet1['AA3']=overallrank[5]
	Sheet1['AB3']=overallrank[6]
	Sheet1['AC3']=overallrank[7]
	Sheet1['AD3']=overallrank[8]
	Sheet1['AE3']=octname(overallrank[8])

	for x in range(3):
	    for c in Sheet1[f'M{x+2}':f'AE{x+2}'][0]:
	        c.border = box



	# --------NOTE---------
	#
	# the final element in the list of rowwise rank is the max
	# element thingy
	#
	# ---------------------



	modstring='Mod '+str(mod)
	Sheet1['M4']=modstring

	#okay time for the mod table ughhh
	ran=(L//mod)+1
	i=0
	j=1

	countrankone=[0,0,0,0,0,0,0,0]

	while i<ran:

		if i==ran-1:
			low=str(j)
			mid='-'
			j=L
			up=str(j)
			string=low+mid+up
			#print (string)
			i=i+1


		else:
			low=str(j)
			mid='-'
			j=j+mod-1
			up=str(j)
			j=j+1
			string=low+mid+up
			#print (string)
			i=i+1

		ind=i+4   #very important! we use this to know where the mod table ends
		sheetiteratorM='M'+str(ind)
		sheetiteratorN='N'+str(ind)
		sheetiteratorO='O'+str(ind)
		sheetiteratorP='P'+str(ind)
		sheetiteratorQ='Q'+str(ind)
		sheetiteratorR='R'+str(ind)
		sheetiteratorS='S'+str(ind)
		sheetiteratorT='T'+str(ind)
		sheetiteratorU='U'+str(ind)
		sheetiteratorV='V'+str(ind)
		sheetiteratorW='W'+str(ind)
		sheetiteratorX='X'+str(ind)
		sheetiteratorY='Y'+str(ind)
		sheetiteratorZ='Z'+str(ind)
		sheetiteratorAA='AA'+str(ind)
		sheetiteratorAB='AB'+str(ind)
		sheetiteratorAC='AC'+str(ind)
		sheetiteratorAD='AD'+str(ind)
		sheetiteratorAE='AE'+str(ind)


		Sheet1[sheetiteratorM]=string

		Sheet1[sheetiteratorN]=cmod[0][i-1]
		Sheet1[sheetiteratorO]=cmod[1][i-1]
		Sheet1[sheetiteratorP]=cmod[2][i-1]
		Sheet1[sheetiteratorQ]=cmod[3][i-1]
		Sheet1[sheetiteratorR]=cmod[4][i-1]
		Sheet1[sheetiteratorS]=cmod[5][i-1]
		Sheet1[sheetiteratorT]=cmod[6][i-1]
		Sheet1[sheetiteratorU]=cmod[7][i-1]

		# okay now time for the new addition
		# drumroooollll

		cmlist=[cmod[0][i-1], cmod[1][i-1], cmod[2][i-1], cmod[3][i-1], cmod[4][i-1], cmod[5][i-1], cmod[6][i-1], cmod[7][i-1]]

		overallrank=[]

		maxvalueranker(cmlist, overallrank)


		Sheet1[sheetiteratorV]=overallrank[0]
		Sheet1[sheetiteratorW]=overallrank[1]
		Sheet1[sheetiteratorX]=overallrank[2]
		Sheet1[sheetiteratorY]=overallrank[3]
		Sheet1[sheetiteratorZ]=overallrank[4]
		Sheet1[sheetiteratorAA]=overallrank[5]
		Sheet1[sheetiteratorAB]=overallrank[6]
		Sheet1[sheetiteratorAC]=overallrank[7]
		Sheet1[sheetiteratorAD]=overallrank[8]
		Sheet1[sheetiteratorAE]=octname(overallrank[8])

		# and now adding up the count for the rank one thing

		if overallrank[8]==1:
			countrankone[0]=countrankone[0]+1

		elif overallrank[8]==-1:
			countrankone[1]=countrankone[1]+1

		elif overallrank[8]==2:
			countrankone[2]=countrankone[2]+1

		elif overallrank[8]==-2:
			countrankone[3]=countrankone[3]+1

		elif overallrank[8]==3:
			countrankone[4]=countrankone[4]+1

		elif overallrank[8]==-3:
			countrankone[5]=countrankone[5]+1

		elif overallrank[8]==4:
			countrankone[6]=countrankone[6]+1

		elif overallrank[8]==-4:
			countrankone[7]=countrankone[7]+1



		# borders!!

		for c in Sheet1[f'M{ind}':f'AE{ind}'][0]:
		    c.border = box





	ind=ind+5

	# print(ind)


	# okay so if you are reading this then just know that i am stupidly
	# idiotic because for some weird reason i was trying to do this
	# individually, cellwise.

	sheetiteratorN='N'+str(ind-1)
	sheetiteratorO='O'+str(ind-1)
	sheetiteratorP='P'+str(ind-1)

	Sheet1[sheetiteratorN]='Octant ID'
	Sheet1[sheetiteratorO]='Octant Name'
	Sheet1[sheetiteratorP]='Count of Rank 1 Mod Values'

	for c in Sheet1[f'N{ind-1}':f'P{ind-1}'][0]:
		c.border = box


	numlist=[1,-1,2,-2,3,-3,4,-4]

	for i in range(8):

		sheetiteratorN='N'+str(ind)
		sheetiteratorO='O'+str(ind)
		sheetiteratorP='P'+str(ind)

		Sheet1[sheetiteratorN]=numlist[i]
		Sheet1[sheetiteratorO]=octname(numlist[i])
		Sheet1[sheetiteratorP]=countrankone[i]

		for c in Sheet1[f'N{ind}':f'P{ind}'][0]:
		    c.border = box


		ind=ind+1

		# now THAT was elegant (something which the rest of my code isnt)






	# we do the overall transcount table here in this bloc

	# ----------START OF ONE BIG BLOC------------

	ind=4

	i=0
	k=ind

	for x in range(9):
	    for c in Sheet1[f'AI{ind-1+x}':f'AQ{ind-1+x}'][0]:
	        c.border = box



	for i in range (8):

		sheetiteratorN='AJ'+str(k)
		sheetiteratorO='AK'+str(k)
		sheetiteratorP='AL'+str(k)
		sheetiteratorQ='AM'+str(k)
		sheetiteratorR='AN'+str(k)
		sheetiteratorS='AO'+str(k)
		sheetiteratorT='AP'+str(k)
		sheetiteratorU='AQ'+str(k)



		Sheet1[sheetiteratorN]=overall_transcount_matrix[i][0]
		Sheet1[sheetiteratorO]=overall_transcount_matrix[i][1]
		Sheet1[sheetiteratorP]=overall_transcount_matrix[i][2]
		Sheet1[sheetiteratorQ]=overall_transcount_matrix[i][3]
		Sheet1[sheetiteratorR]=overall_transcount_matrix[i][4]
		Sheet1[sheetiteratorS]=overall_transcount_matrix[i][5]
		Sheet1[sheetiteratorT]=overall_transcount_matrix[i][6]
		Sheet1[sheetiteratorU]=overall_transcount_matrix[i][7]


		k=k+1



	#THESE DO THE ROW ABOVE THE MATRIX

	pls=ind-1
	plsyameteN='AJ'+str(pls)
	plsyameteO='AK'+str(pls)
	plsyameteP='AL'+str(pls)
	plsyameteQ='AM'+str(pls)
	plsyameteR='AN'+str(pls)
	plsyameteS='AO'+str(pls)
	plsyameteT='AP'+str(pls)
	plsyameteU='AQ'+str(pls)

	Sheet1[plsyameteN]='+1'
	Sheet1[plsyameteO]='-1'
	Sheet1[plsyameteP]='+2'
	Sheet1[plsyameteQ]='-2'
	Sheet1[plsyameteR]='+3'
	Sheet1[plsyameteS]='-3'
	Sheet1[plsyameteT]='+4'
	Sheet1[plsyameteU]='-4'




	#THESE PRINT THE COLUMN TO THE LEFT OF THE MATRIX

	plsyameteMM='AI'+str(ind-2)
	plsyameteM='AI'+str(ind-1)
	plsyameteN='AI'+str(ind)
	plsyameteO='AI'+str(ind+1)
	plsyameteP='AI'+str(ind+2)
	plsyameteQ='AI'+str(ind+3)
	plsyameteR='AI'+str(ind+4)
	plsyameteS='AI'+str(ind+5)
	plsyameteT='AI'+str(ind+6)
	plsyameteU='AI'+str(ind+7)



	string='all the oct values'
	Sheet1[plsyameteMM]=string
	Sheet1[plsyameteM]='Count'
	Sheet1[plsyameteN]='+1'
	Sheet1[plsyameteO]='-1'
	Sheet1[plsyameteP]='+2'
	Sheet1[plsyameteQ]='-2'
	Sheet1[plsyameteR]='+3'
	Sheet1[plsyameteS]='-3'
	Sheet1[plsyameteT]='+4'
	Sheet1[plsyameteU]='-4'

	ind=ind+14



	# ----------END OF ONE BIG BLOC------------




	j=0
	I=0
	J=1
	for j in range (ran):

	    # we use that big bloc again and again for all the matrices we stored
	    # in modtransmatrixlist




	    # ----------START OF ONE BIG BLOC------------

		i=0
		k=ind

		for x in range(9):
		    for c in Sheet1[f'AI{ind-1+x}':f'AQ{ind-1+x}'][0]:
		        c.border = box

		for i in range (8):


			sheetiteratorN='AJ'+str(k)
			sheetiteratorO='AK'+str(k)
			sheetiteratorP='AL'+str(k)
			sheetiteratorQ='AM'+str(k)
			sheetiteratorR='AN'+str(k)
			sheetiteratorS='AO'+str(k)
			sheetiteratorT='AP'+str(k)
			sheetiteratorU='AQ'+str(k)



			Sheet1[sheetiteratorN]=modtransmatrixlist[j][i][0]
			Sheet1[sheetiteratorO]=modtransmatrixlist[j][i][1]
			Sheet1[sheetiteratorP]=modtransmatrixlist[j][i][2]
			Sheet1[sheetiteratorQ]=modtransmatrixlist[j][i][3]
			Sheet1[sheetiteratorR]=modtransmatrixlist[j][i][4]
			Sheet1[sheetiteratorS]=modtransmatrixlist[j][i][5]
			Sheet1[sheetiteratorT]=modtransmatrixlist[j][i][6]
			Sheet1[sheetiteratorU]=modtransmatrixlist[j][i][7]

			k=k+1



	    #THESE DO THE ROW ABOVE THE MATRIX

		pls=ind-1
		plsyameteN='AJ'+str(pls)
		plsyameteO='AK'+str(pls)
		plsyameteP='AL'+str(pls)
		plsyameteQ='AM'+str(pls)
		plsyameteR='AN'+str(pls)
		plsyameteS='AO'+str(pls)
		plsyameteT='AP'+str(pls)
		plsyameteU='AQ'+str(pls)

		Sheet1[plsyameteN]='+1'
		Sheet1[plsyameteO]='-1'
		Sheet1[plsyameteP]='+2'
		Sheet1[plsyameteQ]='-2'
		Sheet1[plsyameteR]='+3'
		Sheet1[plsyameteS]='-3'
		Sheet1[plsyameteT]='+4'
		Sheet1[plsyameteU]='-4'




	    #THESE PRINT THE COLUMN TO THE LEFT OF THE MATRIX

		plsyameteMM='AI'+str(ind-2)
		plsyameteM='AI'+str(ind-1)
		plsyameteN='AI'+str(ind)
		plsyameteO='AI'+str(ind+1)
		plsyameteP='AI'+str(ind+2)
		plsyameteQ='AI'+str(ind+3)
		plsyameteR='AI'+str(ind+4)
		plsyameteS='AI'+str(ind+5)
		plsyameteT='AI'+str(ind+6)
		plsyameteU='AI'+str(ind+7)





		if I==ran-1:
			low=str(J)
			mid='-'
			J=L
			up=str(J)
			string=low+mid+up
			#print (string)
			I=I+1


		else:
			low=str(J)
			mid='-'
			J=J+mod-1
			up=str(J)
			J=J+1
			string=low+mid+up
			#print (string)
			I=I+1

		#string='all the oct values'
		Sheet1[plsyameteMM]=string
		Sheet1[plsyameteM]='Count'
		Sheet1[plsyameteN]='+1'
		Sheet1[plsyameteO]='-1'
		Sheet1[plsyameteP]='+2'
		Sheet1[plsyameteQ]='-2'
		Sheet1[plsyameteR]='+3'
		Sheet1[plsyameteS]='-3'
		Sheet1[plsyameteT]='+4'
		Sheet1[plsyameteU]='-4'

		ind=ind+14



	    # ----------END OF ONE BIG BLOC------------




	# time for the new addition of the table.

	Sheet1['AW3']='Count'
	Sheet1['AX3']='Longest Subsequence Length'
	Sheet1['AY3']='Count'

	for c in Sheet1['AW3:AY3'][0]:
		c.border = box


	row=4
	i=0
	j=0

	# i just realised that it is gonna be kinda difficult with loops
	# given that we've got columns like +1, -1, etc

	# i hate my life

	for i in range (4):


		# -------FOR +VE OCTS--------

		for x in range(3):
		    for c in Sheet1[f'AW{row+x}':f'AY{row+x}'][0]:
		        c.border = box

		iterator_q='AW'+str(row)
		iterator_r='AX'+str(row)
		iterator_s='AY'+str(row)

		num='+'+str(i+1)

		Sheet1[iterator_q]=num
		a=2*i
		Sheet1[iterator_r]=seqlength[a]
		Sheet1[iterator_s]=occount[a]

		row=row+1

		iterator_q='AW'+str(row)
		iterator_r='AX'+str(row)
		iterator_s='AY'+str(row)

		Sheet1[iterator_q]='Time'
		Sheet1[iterator_r]='From'
		Sheet1[iterator_s]='To'



		row=row+1


		j=0
		while j < 0.5*len(timecounter[2*i]):
			iterator_r='AX'+str(row)
			iterator_s='AY'+str(row)

			print(2*i)
			print(2*j)
			print(timecounter[2*i])
			print('+ has run')
			print('---------------')

			Sheet1[iterator_r]=timecounter[2*i][2*j]
			Sheet1[iterator_s]=timecounter[2*i][(2*j)+1]

			for c in Sheet1[f'AW{row}':f'AY{row}'][0]:
			    c.border = box

			row=row+1
			j=j+1


		# -------FOR -VE OCTS--------

		for x in range(3):
		    for c in Sheet1[f'AW{row+x}':f'AY{row+x}'][0]:
		        c.border = box

		print ('it came down')

		iterator_q='AW'+str(row)
		iterator_r='AX'+str(row)
		iterator_s='AY'+str(row)

		num='-'+str(i+1)

		Sheet1[iterator_q]=num
		a=(2*i)+1
		Sheet1[iterator_r]=seqlength[a]
		Sheet1[iterator_s]=occount[a]

		row=row+1

		iterator_q='AW'+str(row)
		iterator_r='AX'+str(row)
		iterator_s='AY'+str(row)

		Sheet1[iterator_q]='Time'
		Sheet1[iterator_r]='From'
		Sheet1[iterator_s]='To'

		row=row+1

		j=0
		while j < 0.5*len(timecounter[(2*i)+1]):
			iterator_r='AX'+str(row)
			iterator_s='AY'+str(row)

			print((2*i)+1)
			print(2*j)
			print(timecounter[(2*i)+1])
			print('- has run')
			print('---------------')

			Sheet1[iterator_r]=timecounter[(2*i)+1][2*j]
			Sheet1[iterator_s]=timecounter[(2*i)+1][(2*j)+1]

			for c in Sheet1[f'AW{row}':f'AY{row}'][0]:
			    c.border = box

			row=row+1
			j=j+1




	# outputfilename='tut07/output/'+inputfile+'_vel_octant_analysis_mod_5000.xlsx'

	buttonlabel='download file ' + str(buttonnumber)


	try:

		if option == 1:



			# -----------------------------DO NOT TOUCH------------------------------------
			output=BytesIO()
			wb2.save(output)
			datafile=output.getvalue()
			infnm=inputfilename.name
			optfnm=infnm[0:7] + '.xlsx'

			st.download_button(label = buttonlabel, data=datafile ,file_name=optfnm)

			# shouldn't have touched it
			# -----------------------------------------------------------------------------


			# wb2.save(outputfilename)

		elif option == 2:
			# -----------------------------DO NOT TOUCH------------------------------------
			output=BytesIO()
			wb2.save(output)
			datafile=output.getvalue()
			infnm=inputfilename[pathlength:]
			optfnm=infnm[0:8] + '.xlsx'

			st.download_button(label = buttonlabel, data=datafile ,file_name=optfnm)

			# shouldn't have touched it
			# -----------------------------------------------------------------------------



	except:
		print("cant save the file")
		print("terminating, adios")
		exit()

	#print(overall_transcount_matrix)
	print('done')



# -----------------------------------------
# MAIN PROGRAM STARTS HERE
# -----------------------------------------



# inputfile1='1.0'
# inputfile2='2.0'
# inputfile3='2.5'
# inputfile2='2.0'
# inputfile2='2.0'
# inputfile2='2.0'
# inputfile2='2.0'
# inputfile2='2.0'
# inputfile2='2.0'
# inputfile2='2.0'
#
# # print('beforecall')
# mainprog(inputfile1)
# # print('aftercall')


# filelist=['1.0', '2.0', '2.5', '3.0', '3.4', '3.7', '4.0', '4.2', '4.4', '4.6', '4.7', '4.8', '4.9', '5.0', '5.1', '5.2', '5.3']

# for inputfile in filelist:
# 	mainprog(inputfile)

st.set_page_config(
	page_title='~ara',
	page_icon=':eggplant:'
	)


selected=option_menu(
	menu_title=None,
	options=['file upload', 'path'],
    icons=['file-earmark-arrow-up', 'signpost-split'],
	default_index=0,
	orientation='horizontal'
)

if selected == 'file upload':

	option=1


	st.title('stonks')

	image = Image.open('stonks.png')

	st.image(image, caption='an accurate representation of our faces when we completed this')

	st.subheader('give it to me senpai~')
	pathlength=1

	inputfiles=st.file_uploader('yes dragging and dropping an entire folder works too', accept_multiple_files=True)
	modval=st.number_input('enter the mod value')
	moddval=int(modval)

	if st.checkbox('execute'):
		if inputfiles:
			buttonnumber=1
			# lol

			for inputfile in inputfiles:
				mainprog(inputfile, moddval, buttonnumber, option, pathlength)
				buttonnumber=buttonnumber+1
		elif not inputfiles:
			st.write('lmao input all values atleast lmao')



		# # -----------------------------DO NOT TOUCH------------------------------------
		# output=BytesIO()
		# wb2.save(output)
		# datafile=output.getvalue()
		# st.download_button(label='📥 Download Current Result', data=datafile ,file_name= 'ara_ara.xlsx')
		# # -----------------------------------------------------------------------------


		# with NamedTemporaryFile() as tmp:
		# 	wb2.save(tmp.name)
		# 	output = BytesIO(tmp.read())
		# # wb2.save()



		#This shall be the last lines of the code.
		# end_time = datetime.now()
		# print('Duration of Program Execution: {}'.format(end_time - start_time))


if selected == 'path':

	option=2

	st.title('path stonks')
	st.subheader('alternatively, you can type in the path of the folder over here')

	inputfiles=st.text_input('enter the path of the folder')
	pathlength=len(inputfiles)
	modval=st.number_input('enter the mod value')
	moddval=int(modval)



	if st.checkbox('execute'):
		if inputfiles:
			buttonnumber=1

			files=os.listdir(inputfiles)

			for file in files:
				inputfile=str(os.path.join(inputfiles, file))
				mainprog(inputfile, moddval, buttonnumber, option, pathlength)
				buttonnumber=buttonnumber+1


		elif not inputfiles:
			st.write('bruh enter the stuff atleast')

