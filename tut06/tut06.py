from platform import python_version
import csv

import pandas as pd

from datetime import datetime
import openpyxl
start_time = datetime.now()


def attendance_report():
    try:
        inp_file = pd.read_csv('input_attendance.csv')
        inp = inp_file.fillna("2001CCXX Random")
    except:
        print("File not found")

    try:
        rollno_inp = pd.read_csv('input_registered_students.csv')
    except:
        print('File containing name of all students is missing!')

    # mr = sheet_input.max_row

    mc = sum(1 for row in open("input_registered_students.csv"))
    mc_consolidated = sum(1 for row in open("input_attendance.csv"))
    total_dates = list()
    for j in range(0, mc_consolidated-1):
        day = inp.at[j, 'Timestamp'].split()[0].split('-')[0]
        month = inp.at[j, 'Timestamp'].split()[0].split('-')[1]
        year = inp.at[j, 'Timestamp'].split()[0].split('-')[2]
        date = datetime.strptime(f'{year}-{month}-{day}', "%Y-%m-%d").date()
        day_name = date.strftime("%A")
        if day_name == "Monday" or day_name == "Thursday":
            if inp.at[j, 'Timestamp'].split()[0] not in total_dates:
                total_dates.append(inp.at[j, 'Timestamp'].split()[0])
    # max_att=0
    # max_roll=""
    fileName_consolidated = ".\output\\attendance_report_consolidated.xlsx"

    output_file = openpyxl.Workbook()
    output = output_file.active
    output_file.save(fileName_consolidated)

    attend_consolidated = pd.read_excel(fileName_consolidated)
    for i in range(0, mc-1):
        total_Present = 0
        date_index = 1
        rollno = rollno_inp.at[i, 'Roll No']
        fileName = ".\output\\"+rollno+'.xlsx'
        output_file = openpyxl.Workbook()
        output = output_file.active
        output_file.save(fileName)
        out = pd.read_excel(fileName)
        for sp_date in total_dates:
            duplicated_attendance = 0
            t_lec, t_lec_act, t_lec_fake, t_lec_abs, percent = len(
                total_dates), 0, 0, 0, 0
            t_lec_count = 0
            for j in range(0, mc_consolidated-1):
                if inp.at[j, 'Attendance'].split()[0] == rollno:
                    if inp.at[j, 'Timestamp'].split()[0] == sp_date:
                        t_lec_count += 1
                        time = inp.at[j, 'Timestamp'].split()[1]
                        hour = time.split(':')[0]
                        minutes = time.split(':')[1]
                        if ((hour == '14') or (hour == '15' and minutes == '00')):
                            if t_lec_act == 0:
                                t_lec_act += 1
                            else:
                                duplicated_attendance += 1
                        else:
                            t_lec_fake += 1

            out.at[0, 'Roll'] = rollno
            out.at[0, 'Name'] = rollno_inp.at[i, 'Name']
            out.at[date_index, 'Dates'] = sp_date
            attend_consolidated.at[i+1, 'Roll'] = rollno
            attend_consolidated.at[i+1, 'Name'] = rollno_inp.at[i, 'Name']
            out.at[date_index, 'Total Attendance Count'] = t_lec_count
            attend_consolidated.at[i+1,
                                   f'{sp_date}'] = 'P' if t_lec_act > 0 else 'A'
            if t_lec_act > 0:
                total_Present += 1
            out.at[date_index, 'Real'] = t_lec_act
            out.at[date_index, 'Duplicate'] = duplicated_attendance
            out.at[date_index, 'Invalid'] = t_lec_fake
            out.at[date_index, 'Absent'] = 1 if t_lec_act == 0 else 0
            date_index += 1
        attend_consolidated.at[i+1, 'Actual Lecture Taken'] = len(total_dates)
        attend_consolidated.at[i+1, 'Total Real'] = total_Present
        attend_consolidated.at[i+1, '% Attendance'] = round(
            total_Present/len(total_dates)*100, 2)
        out.to_excel(fileName, index=False)
        if i == int(0.2*mc):
            print("20% /complete")
        if i == int(0.4*mc):
            print("40% /complete")
        if i == int(0.6*mc):
            print("60% /complete")
        if i == int(0.8*mc):
            print("80% /complete")
        if i == int(mc-2):
            print("100% /complete")

    attend_consolidated.to_excel(fileName_consolidated, index=False)
# ,,,,,, (attendance_count_actual/total_lecture_taken) 2 digit decimal


ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


attendance_report()


# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
